import io
import json
from pathlib import Path
from copy import deepcopy
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

import defaults
import engine
try:
    from openai import OpenAI
except Exception:  # pragma: no cover - optional
    OpenAI = None


BAIN_COLORS = {
    "red": "#CC0000",
    "dark": "#333333",
    "mid": "#666666",
    "light": "#F2F2F2",
    "white": "#FFFFFF",
    "green": "#006600",
}
PALETTE = ["#CC0000", "#333333", "#999999", "#E6E6E6"]


SCENARIO_VARIABLES = [
    {
        "key": "duration",
        "label": "Duration",
        "unit": "days",
        "help": "How long the disruption lasts.",
        "min": 0,
        "type": int,
    },
    {
        "key": "brent",
        "label": "Brent crude",
        "unit": "USD/bbl",
        "help": "Brent benchmark price.",
        "min": 0,
    },
    {
        "key": "jkm",
        "label": "JKM LNG",
        "unit": "USD/MMBtu",
        "help": "Japan-Korea Marker spot LNG price.",
        "min": 0,
    },
    {
        "key": "coal",
        "label": "Newcastle coal",
        "unit": "USD/t",
        "help": "Newcastle thermal coal FOB benchmark.",
        "min": 0,
    },
    {
        "key": "freight",
        "label": "Coal freight",
        "unit": "USD/t",
        "help": "Freight cost from Newcastle to Philippines.",
        "min": 0,
    },
    {
        "key": "insurance",
        "label": "Coal insurance",
        "unit": "USD/t",
        "help": "Marine cargo and war-risk insurance premium per tonne.",
        "min": 0,
    },
    {
        "key": "logistics",
        "label": "Coal logistics adder",
        "unit": "USD/t",
        "help": "Additional logistics disruption cost.",
        "min": 0,
    },
    {
        "key": "fx",
        "label": "FX rate",
        "unit": "PHP/USD",
        "help": "Peso to USD exchange rate.",
        "min": 0.0001,
    },
    {
        "key": "des_basis",
        "label": "LNG DES basis",
        "unit": "USD/MMBtu",
        "help": "Delivered ex-ship basis premium vs JKM.",
        "min": 0,
    },
    {
        "key": "lng_availability",
        "label": "LNG availability",
        "unit": "fraction",
        "help": "Fraction of gas plant capacity that can dispatch.",
        "min": 0,
        "max": 1,
    },
    {
        "key": "coal_delay",
        "label": "Coal delay",
        "unit": "days",
        "help": "Additional days of coal delivery delay.",
        "min": 0,
        "type": int,
    },
    {
        "key": "bid_uplift",
        "label": "Bid uplift",
        "unit": "fraction",
        "help": "WESM bid mark-up above marginal cost.",
        "min": 0,
        "max": 1,
    },
    {
        "key": "reserve_alpha",
        "label": "Reserve alpha",
        "unit": "fraction",
        "help": "Fraction of WESM price used as reserve anchor.",
        "min": 0,
        "max": 1,
    },
    {
        "key": "bad_debt_multiplier",
        "label": "Bad debt multiplier",
        "unit": "x",
        "help": "Multiplier on retail base bad-debt rates.",
        "min": 1,
    },
]


def inject_css():
    st.markdown(
        f"""
        <style>
        html, body, [class*="block-container"] {{
            font-family: Arial, Helvetica, sans-serif !important;
            background: {BAIN_COLORS['white']};
        }}
        .stApp {{
            --primary-color: {BAIN_COLORS['red']};
            --text-color: {BAIN_COLORS['dark']};
            --secondary-text-color: {BAIN_COLORS['mid']};
            --background-color: {BAIN_COLORS['white']};
            --surface-color: {BAIN_COLORS['light']};
            --border-radius: 0px;
        }}
        .bain-hr {{
            border: 0;
            border-top: 2px solid {BAIN_COLORS['red']};
            margin: 0 0 12px 0;
        }}
        .bain-card {{
            border: 1px solid #ccc;
            padding: 12px;
            background: {BAIN_COLORS['white']};
            position: relative;
        }}
        .bain-card::before {{
            content: "";
            position: absolute;
            left: 0;
            top: 0;
            bottom: 0;
            width: 4px;
            background: {BAIN_COLORS['red']};
        }}
        .bain-table-header {{
            background: {BAIN_COLORS['dark']} !important;
            color: white !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def init_state():
    if "scenarios" not in st.session_state:
        st.session_state.scenarios = deepcopy(defaults.SCENARIOS_DEFAULT)
    if "scenario_names" not in st.session_state:
        st.session_state.scenario_names = list(defaults.SCENARIOS_DEFAULT["names"])
    if "api_key" not in st.session_state:
        st.session_state.api_key = ""
    if "coal_quality_key" not in st.session_state:
        st.session_state.coal_quality_key = "GAR 5250"
    if "season_key" not in st.session_state:
        st.session_state.season_key = "Hot-dry"
    if "spc_toggle" not in st.session_state:
        st.session_state.spc_toggle = defaults.SPC["toggle"]
        st.session_state.spc_cap = defaults.SPC["cap"]

    # Tab 2 defaults cached as dataframes for easy editing
    if "plants_df" not in st.session_state:
        st.session_state.plants_df = pd.DataFrame(
            [
                ["Coal cluster", "Coal", defaults.PLANTS["coal"]["capacity"], defaults.PLANTS["coal"]["availability"],
                 defaults.PLANTS["coal"]["heat_rate"], defaults.PLANTS["coal"]["vom"], defaults.PLANTS["coal"]["base_dispatch"]],
                ["Gas cluster", "Gas/LNG", defaults.PLANTS["gas"]["capacity"], defaults.PLANTS["gas"]["availability"],
                 defaults.PLANTS["gas"]["heat_rate"], defaults.PLANTS["gas"]["vom"], defaults.PLANTS["gas"]["base_dispatch"]],
                ["Oil peakers", "Oil", defaults.PLANTS["oil"]["capacity"], defaults.PLANTS["oil"]["availability"],
                 defaults.PLANTS["oil"]["heat_rate"], defaults.PLANTS["oil"]["vom"], defaults.PLANTS["oil"]["base_dispatch"]],
            ],
            columns=[
                "Cluster",
                "Fuel",
                "Capacity (MW)",
                "Availability (%)",
                "Heat Rate (GJ/MWh)",
                "Variable O&M (PHP/MWh)",
                "Base Dispatch (MWh/month)",
            ],
        )

    if "system_df" not in st.session_state:
        rows = []
        for c in defaults.SYSTEM_STACK:
            srmc_list = c.get("srmc_base") or [None, None, None, None]
            rows.append(
                [
                    c["name"],
                    c["dep_mw"],
                    c["tech_avail"],
                    c["profile"]["off_peak"],
                    c["profile"]["shoulder"],
                    c["profile"]["peak"],
                    srmc_list[0],
                    srmc_list[1],
                    srmc_list[2],
                    srmc_list[3],
                ]
            )
        st.session_state.system_df = pd.DataFrame(
            rows,
            columns=[
                "Cluster",
                "Dep. MW",
                "Tech Avail %",
                "Off-peak profile",
                "Shoulder profile",
                "Peak profile",
                "SRMC Base",
                "SRMC S1",
                "SRMC S2",
                "SRMC S3",
            ],
        )

    if "lng_df" not in st.session_state:
        st.session_state.lng_df = pd.DataFrame(
            {
                "Component": ["Regas / terminal fee", "Delay / demurrage"],
                "Base": [defaults.FUEL_INTERMEDIATES["regas"][0], defaults.FUEL_INTERMEDIATES["demurrage"][0]],
                "S1": [defaults.FUEL_INTERMEDIATES["regas"][1], defaults.FUEL_INTERMEDIATES["demurrage"][1]],
                "S2": [defaults.FUEL_INTERMEDIATES["regas"][2], defaults.FUEL_INTERMEDIATES["demurrage"][2]],
                "S3": [defaults.FUEL_INTERMEDIATES["regas"][3], defaults.FUEL_INTERMEDIATES["demurrage"][3]],
            }
        )

    if "contracts_df" not in st.session_state:
        st.session_state.contracts_df = pd.DataFrame(
            {
                "Input": [
                    "BCQ volume",
                    "BCQ sell price",
                    "Spot sales volume",
                    "Imbalance volume",
                    "Imbalance reference price",
                    "Repricing beta",
                    "Wholesale opex",
                ],
                "Value": [
                    defaults.CONTRACTS["bcq_volume"],
                    defaults.CONTRACTS["bcq_price"],
                    defaults.CONTRACTS["spot_volume"],
                    defaults.CONTRACTS["imbalance_volume"],
                    defaults.CONTRACTS["imbalance_ref_price"],
                    defaults.CONTRACTS["repricing_beta"],
                    defaults.CONTRACTS["wholesale_opex"],
                ],
                "Unit": ["MWh/mo", "PHP/MWh", "MWh/mo", "MWh/mo", "PHP/MWh", "fraction", "PHP mn/mo"],
            }
        )

    if "retail_df" not in st.session_state:
        st.session_state.retail_df = pd.DataFrame(defaults.RETAIL)

    if "liquidity_df" not in st.session_state:
        st.session_state.liquidity_df = pd.DataFrame(
            list(defaults.LIQUIDITY.items()), columns=["Input", "Value"]
        )

    if "agra_df" not in st.session_state:
        ir = defaults.AGRA["immediate_recovery"]
        ds = defaults.AGRA["disallowance"]
        cl = defaults.AGRA["collection_lag"]
        rl = defaults.AGRA["refund_lag"]
        tu = defaults.AGRA["true_up_horizon"]
        cc = defaults.AGRA["carrying_cost"]
        st.session_state.agra_df = pd.DataFrame(
            {
                "Parameter": [
                    "Immediate recovery %",
                    "Disallowance %",
                    "Collection lag (days)",
                    "Refund lag (months)",
                    "True-up horizon (months)",
                    "Carrying cost p.a.",
                ],
                "Base": [ir[0], ds[0], cl[0], rl[0], tu[0], cc[0]],
                "30d": [ir[1], ds[1], cl[1], rl[1], tu[1], cc[1]],
                "90d": [ir[2], ds[2], cl[2], rl[2], tu[2], cc[2]],
                "180d": [ir[3], ds[3], cl[3], rl[3], tu[3], cc[3]],
            }
        )

    if "triggers_df" not in st.session_state:
        st.session_state.triggers_df = pd.DataFrame(defaults.TRIGGERS)

    if "demand_df" not in st.session_state:
        st.session_state.demand_df = pd.DataFrame(
            {
                "Block": ["Off-peak", "Shoulder", "Peak"],
                "Base MW": [
                    defaults.DEMAND_BLOCKS["off_peak_base"],
                    defaults.DEMAND_BLOCKS["shoulder_base"],
                    defaults.DEMAND_BLOCKS["peak_base"],
                ],
                "Hours": [
                    defaults.DEMAND_BLOCKS["hours_off"],
                    defaults.DEMAND_BLOCKS["hours_shoulder"],
                    defaults.DEMAND_BLOCKS["hours_peak"],
                ],
                "Description": [
                    "Night trough demand",
                    "Daytime / evening shoulder",
                    "System peak",
                ],
            }
        )

    if "vre_df" not in st.session_state:
        st.session_state.vre_df = pd.DataFrame(
            {
                "Source": ["Solar", "Wind"],
                "Off-peak": [defaults.VRE_SHAPES["solar"]["off_peak"], defaults.VRE_SHAPES["wind"]["off_peak"]],
                "Shoulder": [defaults.VRE_SHAPES["solar"]["shoulder"], defaults.VRE_SHAPES["wind"]["shoulder"]],
                "Peak": [defaults.VRE_SHAPES["solar"]["peak"], defaults.VRE_SHAPES["wind"]["peak"]],
            }
        )

    if "hydro_df" not in st.session_state:
        st.session_state.hydro_df = pd.DataFrame(
            {
                "Setting": ["Hydro budget", "Hydro peak MW limit"],
                "Hot-dry": [16, 1200],
                "Wet": [22, 1500],
                "Cool-dry": [18, 1300],
                "Unit": ["GWh/day", "MW"],
            }
        )

    if "_recalc_clicks" not in st.session_state:
        st.session_state["_recalc_clicks"] = 0


def scenario_df() -> pd.DataFrame:
    data = []
    names = st.session_state.scenario_names
    n = len(names)
    for row in SCENARIO_VARIABLES:
        vals = np.array(st.session_state.scenarios[row["key"]], dtype=float)
        if len(vals) < n:
            vals = np.concatenate([vals, np.full(n - len(vals), vals[-1] if len(vals) else 0)])
        elif len(vals) > n:
            vals = vals[:n]
        data.append([row["label"], row["unit"], *vals])
    return pd.DataFrame(
        data,
        columns=["Variable", "Unit", *names],
    )


def update_scenarios_from_df(df: pd.DataFrame):
    for row in SCENARIO_VARIABLES:
        vals = df.loc[df["Variable"] == row["label"], st.session_state.scenario_names].values[0]
        st.session_state.scenarios[row["key"]] = np.array(vals, dtype=float)


def parse_uploaded_scenarios(uploaded_file):
    try:
        # Prefer pandas for flexible header handling
        df_raw = pd.read_excel(uploaded_file, sheet_name="03_Scenarios")
    except Exception:
        st.warning("Sheet 03_Scenarios not found or unreadable.")
        return

    # Normalize column names
    df_raw.columns = [str(c).strip() for c in df_raw.columns]
    # Find scenario columns (Base + scenarios)
    scenario_cols = [c for c in df_raw.columns if c.lower().startswith("base") or c.lower().startswith("scenario")]
    if len(scenario_cols) < 4:
        st.warning("Expected at least 4 scenario columns (Base, Scenario 1-3).")
        return

    # Preserve scenario names from header
    st.session_state.scenario_names = scenario_cols[:4]

    # Align variables
    for row in SCENARIO_VARIABLES:
        label = row["label"]
        key = row["key"]
        match = df_raw[df_raw[df_raw.columns[0]].astype(str).str.strip().str.lower() == label.lower()]
        if match.empty:
            st.warning(f"Variable '{label}' not found in uploaded sheet.")
            continue
        values = match.iloc[0][scenario_cols[:4]].values
        try:
            st.session_state.scenarios[key] = np.array(values, dtype=float)
        except Exception:
            st.warning(f"Could not parse values for {label}.")

    st.success(f"Loaded {len(SCENARIO_VARIABLES)} variables × 4 scenarios from {uploaded_file.name}")


def build_inputs_from_state():
    spc = {"toggle": int(st.session_state.spc_toggle), "cap": st.session_state.spc_cap}

    # Plants
    plant_df = st.session_state.plants_df.copy()
    plants = {
        "coal": {
            "capacity": plant_df.iloc[0]["Capacity (MW)"],
            "availability": plant_df.iloc[0]["Availability (%)"],
            "heat_rate": plant_df.iloc[0]["Heat Rate (GJ/MWh)"],
            "vom": plant_df.iloc[0]["Variable O&M (PHP/MWh)"],
            "base_dispatch": plant_df.iloc[0]["Base Dispatch (MWh/month)"],
        },
        "gas": {
            "capacity": plant_df.iloc[1]["Capacity (MW)"],
            "availability": plant_df.iloc[1]["Availability (%)"],
            "heat_rate": plant_df.iloc[1]["Heat Rate (GJ/MWh)"],
            "vom": plant_df.iloc[1]["Variable O&M (PHP/MWh)"],
            "base_dispatch": plant_df.iloc[1]["Base Dispatch (MWh/month)"],
        },
        "oil": {
            "capacity": plant_df.iloc[2]["Capacity (MW)"],
            "availability": plant_df.iloc[2]["Availability (%)"],
            "heat_rate": plant_df.iloc[2]["Heat Rate (GJ/MWh)"],
            "vom": plant_df.iloc[2]["Variable O&M (PHP/MWh)"],
            "base_dispatch": plant_df.iloc[2]["Base Dispatch (MWh/month)"],
        },
        "oil_product_diff": defaults.PLANTS["oil_product_diff"],
    }

    # System stack
    sys_df = st.session_state.system_df.copy()
    system_stack = []
    for _, r in sys_df.iterrows():
        profile = {
            "off_peak": r["Off-peak profile"],
            "shoulder": r["Shoulder profile"],
            "peak": r["Peak profile"],
        }
        srmc_base = [r["SRMC Base"], r["SRMC S1"], r["SRMC S2"], r["SRMC S3"]]
        system_stack.append(
            {
                "name": r["Cluster"],
                "dep_mw": r["Dep. MW"],
                "tech_avail": r["Tech Avail %"],
                "profile": profile,
                "srmc_base": None if any(pd.isna(srmc_base)) else srmc_base,
            }
        )

    # LNG fees
    lng_df = st.session_state.lng_df.set_index("Component")
    fuel_intermediates = {
        "regas": lng_df.loc["Regas / terminal fee", ["Base", "S1", "S2", "S3"]].to_numpy(dtype=float),
        "demurrage": lng_df.loc["Delay / demurrage", ["Base", "S1", "S2", "S3"]].to_numpy(dtype=float),
    }

    # Contracts
    cdf = st.session_state.contracts_df.set_index("Input")
    contracts = {
        "bcq_volume": float(cdf.loc["BCQ volume", "Value"]),
        "bcq_price": float(cdf.loc["BCQ sell price", "Value"]),
        "spot_volume": float(cdf.loc["Spot sales volume", "Value"]),
        "imbalance_volume": float(cdf.loc["Imbalance volume", "Value"]),
        "imbalance_ref_price": float(cdf.loc["Imbalance reference price", "Value"]),
        "repricing_beta": float(cdf.loc["Repricing beta", "Value"]),
        "wholesale_opex": float(cdf.loc["Wholesale opex", "Value"]),
    }

    # Retail list
    retail_df = st.session_state.retail_df
    retail = []
    for _, r in retail_df.iterrows():
        retail.append(
            {
                "name": r["name"],
                "volume": r["volume"],
                "base_tariff": r["base_tariff"],
                "wesm_link": r["wesm_link"],
                "pass_through": r["pass_through"],
                "bad_debt_base": r["bad_debt_base"],
                "attrition": r["attrition"],
                "opex": r["opex"],
            }
        )

    import ast
    liquidity_df = st.session_state.liquidity_df.set_index("Input")
    liquidity: Dict[str, object] = {}
    for k in liquidity_df.index:
        raw_val = liquidity_df.loc[k, "Value"]
        if k == "lc_margin":
            if isinstance(raw_val, str):
                try:
                    raw_val = ast.literal_eval(raw_val)
                except Exception:
                    pass
            liquidity[k] = np.array(raw_val, dtype=float)
        else:
            try:
                liquidity[k] = float(raw_val)
            except Exception:
                liquidity[k] = float(str(raw_val).replace(",", ""))

    agra_df = st.session_state.agra_df.set_index("Parameter")
    agra = {
        "immediate_recovery": agra_df.loc["Immediate recovery %", ["Base", "30d", "90d", "180d"]].to_numpy(dtype=float),
        "disallowance": agra_df.loc["Disallowance %", ["Base", "30d", "90d", "180d"]].to_numpy(dtype=float),
        "collection_lag": agra_df.loc["Collection lag (days)", ["Base", "30d", "90d", "180d"]].to_numpy(dtype=float),
        "refund_lag": agra_df.loc["Refund lag (months)", ["Base", "30d", "90d", "180d"]].to_numpy(dtype=float),
        "true_up_horizon": agra_df.loc["True-up horizon (months)", ["Base", "30d", "90d", "180d"]].to_numpy(dtype=float),
        "carrying_cost": agra_df.loc["Carrying cost p.a.", ["Base", "30d", "90d", "180d"]].to_numpy(dtype=float),
    }

    triggers = st.session_state.triggers_df.to_dict(orient="records")

    demand_df = st.session_state.demand_df
    demand_blocks = {
        "off_peak_base": demand_df.iloc[0]["Base MW"],
        "shoulder_base": demand_df.iloc[1]["Base MW"],
        "peak_base": demand_df.iloc[2]["Base MW"],
        "hours_off": demand_df.iloc[0]["Hours"],
        "hours_shoulder": demand_df.iloc[1]["Hours"],
        "hours_peak": demand_df.iloc[2]["Hours"],
    }

    vre_df = st.session_state.vre_df.set_index("Source")
    vre_shapes = {
        "solar": {
            "off_peak": vre_df.loc["Solar", "Off-peak"],
            "shoulder": vre_df.loc["Solar", "Shoulder"],
            "peak": vre_df.loc["Solar", "Peak"],
        },
        "wind": {
            "off_peak": vre_df.loc["Wind", "Off-peak"],
            "shoulder": vre_df.loc["Wind", "Shoulder"],
            "peak": vre_df.loc["Wind", "Peak"],
        },
    }

    hydro_df = st.session_state.hydro_df.set_index("Setting")
    season_table = deepcopy(defaults.SEASONS)
    for season in season_table.keys():
        season_table[season]["hydro_budget_gwh"] = hydro_df.loc["Hydro budget", season]
        season_table[season]["hydro_peak_limit"] = hydro_df.loc["Hydro peak MW limit", season]

    inputs = {
        "scenarios": st.session_state.scenarios,
        "constants": defaults.CONSTANTS,
        "coal_quality": defaults.COAL_QUALITY[st.session_state.coal_quality_key],
        "season": {"selected": st.session_state.season_key, "table": season_table},
        "demand_blocks": demand_blocks,
        "vre_shapes": vre_shapes,
        "spc": spc,
        "plants": plants,
        "system_stack": system_stack,
        "fuel_intermediates": fuel_intermediates,
        "contracts": contracts,
        "retail": retail,
        "liquidity": liquidity,
        "agra": agra,
        "triggers": triggers,
        "reserve": defaults.RESERVE,
    }
    return inputs


def format_money(val, decimals=1):
    try:
        return f"{val:,.{decimals}f}"
    except Exception:
        return "—"


def format_pct(val):
    try:
        return f"{val*100:.1f}%"
    except Exception:
        return "—"


def recalc_button(label: str = "🔄 Recalculate & Refresh Results", key: Optional[str] = None):
    # Provide unique keys to avoid Streamlit duplicate element IDs across tabs
    btn_key = key or label
    if st.button(label, type="primary", use_container_width=True, key=btn_key):
        st.session_state["_recalc_clicks"] += 1
        st.rerun()


def df_to_text(df: pd.DataFrame) -> str:
    """Safe textual table without requiring tabulate."""
    try:
        return df.to_markdown(index=False)
    except Exception:
        return df.to_string(index=False)


def add_scenario(name: str, values: Dict[str, float]):
    if len(st.session_state.scenario_names) >= 5:
        st.warning("Maximum of 5 scenarios reached. Delete one before adding a new scenario.")
        return False
    st.session_state.scenario_names.append(name)
    for row in SCENARIO_VARIABLES:
        key = row["key"]
        current = np.array(st.session_state.scenarios[key], dtype=float)
        st.session_state.scenarios[key] = np.append(current, float(values.get(key, current[-1] if len(current) else 0)))
    return True


def parse_json_safely(text: str):
    """Try to parse JSON from arbitrary LLM text."""
    text = text.strip()
    try:
        return json.loads(text)
    except Exception:
        # try to extract first {...}
        start = text.find("{")
        end = text.rfind("}")
        if start != -1 and end != -1 and end > start:
            snippet = text[start : end + 1]
            try:
                return json.loads(snippet)
            except Exception:
                pass
    return None


def get_openai_client():
    if not OpenAI:
        st.warning("openai package not available. Install requirements.txt.")
        return None
    key = st.session_state.get("api_key") or ""
    if not key:
        st.info("Add an OpenAI API key in the sidebar to enable AI features.")
        return None
    try:
        return OpenAI(api_key=key)
    except Exception as e:
        st.error(f"OpenAI client error: {e}")
        return None


def ai_explain(section_key: str, title: str, context: str):
    client = get_openai_client()
    if client is None:
        return
    btn = st.button(f"✨ AI explain {title}", key=f"ai_{section_key}")
    if not btn:
        return
    try:
        scen_names = ", ".join(st.session_state.get("scenario_names", []))
        prompt = (
            "You are a Bain-style energy & utilities risk advisor. Write a sharp, insight-rich explanation "
            "of the results below. Use 4–6 bullets, ~150–220 words total. "
            "Emphasize what changes across scenarios, why (drivers), and the business implications. "
            "Call out Base vs worst-case, major inflections, liquidity/runway pressure, SRMC vs WESM gaps, "
            "trigger statuses, and any red flags. Use numbers from the context when helpful. "
            f"Scenario names to reference: {scen_names}. "
            "Keep it plain text bullets (start lines with '- '), no tables."
        )
        completion = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": prompt},
                {"role": "user", "content": context[:6000]},
            ],
            temperature=0.35,
            max_tokens=320,
        )
        text = completion.choices[0].message.content.strip()
        st.success(text)
    except Exception as e:
        st.error(f"AI explanation failed: {e}")


def delete_scenario(name: str):
    if name not in st.session_state.scenario_names:
        return
    idx = st.session_state.scenario_names.index(name)
    if idx == 0:
        st.warning("Base scenario cannot be deleted.")
        return
    st.session_state.scenario_names.pop(idx)
    for row in SCENARIO_VARIABLES:
        key = row["key"]
        arr = list(st.session_state.scenarios[key])
        if idx < len(arr):
            arr.pop(idx)
            st.session_state.scenarios[key] = np.array(arr, dtype=float)


def render_kpi(label: str, value, unit: str = "", positive_is_good: Optional[bool] = None):
    color = BAIN_COLORS["dark"]
    if positive_is_good is not None:
        if (value or 0) >= 0 and positive_is_good:
            color = BAIN_COLORS["green"]
        if (value or 0) < 0 and positive_is_good:
            color = BAIN_COLORS["red"]
        if (value or 0) >= 0 and not positive_is_good:
            color = BAIN_COLORS["red"]
        if (value or 0) < 0 and not positive_is_good:
            color = BAIN_COLORS["green"]
    st.markdown(
        f"""
        <div class="bain-card">
            <div style="font-size:12px;color:{BAIN_COLORS['mid']};">{label}</div>
            <div style="font-size:24px;color:{color};font-weight:600;">{format_money(value)} {unit}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def chart_srmc(results, scenario_names):
    df = pd.DataFrame(
        {
            "Scenario": scenario_names,
            "Coal": results["plant_srmc"]["coal_srmc"],
            "Gas": results["plant_srmc"]["gas_srmc"],
            "Oil": results["plant_srmc"]["oil_srmc"],
        }
    )
    df = df.melt(id_vars="Scenario", var_name="Fuel", value_name="SRMC")
    fig = px.bar(df, x="Scenario", y="SRMC", color="Fuel", barmode="group", color_discrete_sequence=PALETTE)
    fig.update_layout(showlegend=True, template="simple_white", yaxis_title="PHP/MWh", xaxis_title="")
    fig.update_yaxes(gridcolor="rgba(0,0,0,0)")
    return fig


def chart_wesm(results, scenario_names):
    df = pd.DataFrame(
        {
            "Scenario": scenario_names,
            "WESM": results["merit_order"]["wesm_avg"],
        }
    )
    fig = px.line(df, x="Scenario", y="WESM", markers=True, color_discrete_sequence=[BAIN_COLORS["red"]])
    fig.update_layout(template="simple_white", yaxis_title="PHP/MWh", xaxis_title="")
    return fig


def chart_retail(results, scenario_names):
    segs = results["retail"]["segments"]
    data = []
    for seg in segs:
        for i, sc in enumerate(scenario_names):
            data.append({"Scenario": sc, "Segment": seg["name"], "EBIT": seg["ebit"][i]})
    df = pd.DataFrame(data)
    fig = px.bar(df, x="Scenario", y="EBIT", color="Segment", barmode="stack", color_discrete_sequence=PALETTE)
    fig.update_layout(template="simple_white", yaxis_title="PHP mn/mo", xaxis_title="")
    return fig


def chart_liquidity(results, scenario_names):
    liq = results["liquidity"]
    df = pd.DataFrame(
        {
            "Scenario": scenario_names,
            "Working Capital": liq["working_capital"],
            "LC": liq["lc_req"],
            "Reg Receivable": liq["reg_receivable"],
            "Refund": liq["refund_provision"],
        }
    )
    fig = px.bar(
        df,
        x="Scenario",
        y=["Working Capital", "LC", "Reg Receivable", "Refund"],
        barmode="stack",
        color_discrete_sequence=PALETTE,
    )
    fig.update_layout(template="simple_white", yaxis_title="PHP mn", xaxis_title="")
    return fig


def download_results_xlsx(results, scenario_names):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "60_Outputs"

    def write_table(start_row, title, headers, rows):
        ws.cell(start_row, 1, title).font = Font(bold=True)
        row = start_row + 1
        ws.append([None])
        ws.append(headers)
        for r in rows:
            ws.append(r)
        # style header
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row + 1, col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        return row + 1 + len(rows) + 1

    r = 1
    headers = ["Variable", "Unit", *scenario_names]
    table_a = []
    for i, row in enumerate(SCENARIO_VARIABLES):
        vals = results["scenarios"][row["key"]]
        unit = row["unit"]
        table_a.append([row["label"], unit, *vals])
    r = write_table(r, "Scenario Matrix", headers, table_a)

    headers_b = ["Metric", "Unit", *scenario_names]
    gm = results["gen_margin"]
    mo = results["merit_order"]
    table_b = [
        ["Coal SRMC", "PHP/MWh", *results["plant_srmc"]["coal_srmc"]],
        ["Gas SRMC", "PHP/MWh", *results["plant_srmc"]["gas_srmc"]],
        ["Oil SRMC", "PHP/MWh", *results["plant_srmc"]["oil_srmc"]],
        ["WESM Avg", "PHP/MWh", *mo["wesm_avg"]],
        ["Reserve Price", "PHP/MWh", *results["reserve_pricing"]["reserve_price"]],
        ["Generation Gross Margin", "PHP mn/mo", *gm["total_gen_margin"]],
    ]
    r = write_table(r + 1, "Generation Sensitivity", headers_b, table_b)

    sw = results["supply_wholesale"]
    ret = results["retail"]
    table_c = [
        ["BCQ Exposure", "PHP mn/mo", *sw["bcq_margin"]],
        ["Spot Sales", "PHP mn/mo", *sw["spot_margin"]],
        ["Portfolio Retail EBIT", "PHP mn/mo", *ret["portfolio_ebit"]],
        ["Supply & Wholesale Margin", "PHP mn/mo", *sw["total_sw_margin"]],
    ]
    r = write_table(r + 1, "Supply & Retail", headers_b, table_c)

    liq = results["liquidity"]
    headers_d = ["Metric", "Unit", *scenario_names]
    table_d = [
        ["Working Capital", "PHP mn", *liq["working_capital"]],
        ["LC Requirements", "PHP mn", *liq["lc_req"]],
        ["Cash Balance", "PHP mn", *liq["cash_balance"]],
        ["Covenant Headroom", "PHP mn", *liq["covenant_headroom"]],
        ["Liquidity Runway", "days", *liq["runway_days"]],
        ["Regulatory Receivable", "PHP mn", *liq["reg_receivable"]],
        ["Refund Provision", "PHP mn", *liq["refund_provision"]],
    ]
    write_table(r + 1, "Liquidity", headers_d, table_d)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def results_tables(results, scenario_names):
    gm = results["gen_margin"]
    mo = results["merit_order"]
    sw = results["supply_wholesale"]
    ret = results["retail"]
    liq = results["liquidity"]

    table_b = pd.DataFrame(
        {
            "Metric": ["Coal SRMC", "Gas SRMC", "Oil SRMC", "WESM Avg", "Reserve Price", "Generation Gross Margin"],
            "Unit": ["PHP/MWh"] * 5 + ["PHP mn/mo"],
            scenario_names[0]: [
                *[results["plant_srmc"]["coal_srmc"][0], results["plant_srmc"]["gas_srmc"][0], results["plant_srmc"]["oil_srmc"][0], mo["wesm_avg"][0], results["reserve_pricing"]["reserve_price"][0]],
                gm["total_gen_margin"][0],
            ],
            scenario_names[1]: [
                *[results["plant_srmc"]["coal_srmc"][1], results["plant_srmc"]["gas_srmc"][1], results["plant_srmc"]["oil_srmc"][1], mo["wesm_avg"][1], results["reserve_pricing"]["reserve_price"][1]],
                gm["total_gen_margin"][1],
            ],
            scenario_names[2]: [
                *[results["plant_srmc"]["coal_srmc"][2], results["plant_srmc"]["gas_srmc"][2], results["plant_srmc"]["oil_srmc"][2], mo["wesm_avg"][2], results["reserve_pricing"]["reserve_price"][2]],
                gm["total_gen_margin"][2],
            ],
            scenario_names[3]: [
                *[results["plant_srmc"]["coal_srmc"][3], results["plant_srmc"]["gas_srmc"][3], results["plant_srmc"]["oil_srmc"][3], mo["wesm_avg"][3], results["reserve_pricing"]["reserve_price"][3]],
                gm["total_gen_margin"][3],
            ],
        }
    )

    table_c = pd.DataFrame(
        {
            "Segment": ["BCQ Exposure", "Spot Sales", "Portfolio Retail EBIT", "Supply & Wholesale Margin"],
            "Unit": ["PHP mn/mo"] * 4,
            scenario_names[0]: [sw["bcq_margin"][0], sw["spot_margin"][0], ret["portfolio_ebit"][0], sw["total_sw_margin"][0]],
            scenario_names[1]: [sw["bcq_margin"][1], sw["spot_margin"][1], ret["portfolio_ebit"][1], sw["total_sw_margin"][1]],
            scenario_names[2]: [sw["bcq_margin"][2], sw["spot_margin"][2], ret["portfolio_ebit"][2], sw["total_sw_margin"][2]],
            scenario_names[3]: [sw["bcq_margin"][3], sw["spot_margin"][3], ret["portfolio_ebit"][3], sw["total_sw_margin"][3]],
        }
    )

    table_d = pd.DataFrame(
        {
            "Metric": ["Working Capital", "LC Requirements", "Cash Balance", "Covenant Headroom", "Liquidity Runway", "Regulatory Receivable", "Refund Provision"],
            "Unit": ["PHP mn", "PHP mn", "PHP mn", "PHP mn", "days", "PHP mn", "PHP mn"],
            scenario_names[0]: liq["working_capital"],
            "30 Days": liq["lc_req"],
            "90 Days": liq["cash_balance"],
            "180 Days": liq["covenant_headroom"],
        }
    )

    return table_b, table_c, liq


def main():
    st.set_page_config(page_title="Gen Co Stress Test", layout="wide")
    inject_css()
    init_state()

    # Sidebar: description, guide, recalc
    st.sidebar.title("Gen Co Stress Test")
    st.sidebar.markdown(
        """
        **What this tool does**
        - Stress-test Gen Co's generation + retail portfolio under commodity, FX, and logistics shocks.
        - Replicates the Bain Excel model logic (fuel landing → SRMC → merit order → margins → liquidity).
        - Compare scenarios side-by-side; export dashboard tables to Excel.

        **How to use**
        1) Define scenarios (manual or upload Excel/news).
        2) Adjust structural inputs (Tab 2).
        3) Hit **Recalculate** to refresh results.
        4) Review dashboard & downloads.
        """
    )
    st.sidebar.markdown("---")
    st.sidebar.text_input("OpenAI API Key (for AI features)", type="password", key="api_key")
    st.sidebar.caption("Key is used locally for news-to-scenario and AI explanations.")
    st.sidebar.markdown("---")
    with st.sidebar:
        recalc_button(key="recalc_sidebar")

    st.title("Gen Co Business Continuity Stress Test")
    st.markdown("<hr class='bain-hr'>", unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs([
        "① Scenario Definition",
        "② Model Inputs",
        "③ Results & Dashboard",
        "④ Intermediate Calculations",
    ])

    # Tab 1 — Scenario Definition
    with tab1:
        st.subheader("Scenario Definition")
        mode = st.radio("Input mode", ["Manual entry", "Upload Excel"], horizontal=True)
        if mode == "Upload Excel":
            st.info("Upload expects one Base plus Scenario 1–3 columns (exactly 4 scenarios).")
            template_path = Path("scenario_template.xlsx")
            if template_path.exists():
                with template_path.open("rb") as f:
                    st.download_button(
                        "⬇️ Download scenario template (.xlsx)",
                        data=f.read(),
                        file_name="scenario_template.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
            uploaded = st.file_uploader("Upload 03_Scenarios range", type=["xlsx"])
            if uploaded:
                parse_uploaded_scenarios(uploaded)

        cols = st.columns(4)
        for i, name in enumerate(["Base", "Scenario 1", "Scenario 2", "Scenario 3"]):
            st.session_state.scenario_names[i] = cols[i].text_input(
                f"Scenario {i}", value=st.session_state.scenario_names[i]
            )

        # Editable grid
        df = scenario_df()
        edited = st.data_editor(df, num_rows="fixed", key="scenario_editor")
        update_scenarios_from_df(edited)
        if len(st.session_state.scenario_names) > 5:
            st.warning("Maximum of 5 scenarios supported. Trimming to first 5.")
            st.session_state.scenario_names = st.session_state.scenario_names[:5]
            for row in SCENARIO_VARIABLES:
                st.session_state.scenarios[row["key"]] = np.array(st.session_state.scenarios[row["key"]][:5])

        st.info(
            "Scenario framework: S1 = Contained disruption / rapid normalisation | "
            "S2 = Sustained regional disruption | S3 = Full-scale regional supply shock"
        )

        st.markdown("**Generate scenario from news (.txt)**")
        news_file = st.file_uploader("Upload concatenated news (.txt)", type=["txt"], key="news_txt")
        new_scen_name = st.text_input("New scenario name", value="News Scenario")
        if st.button("Generate scenario from news", key="btn_news_scenario"):
            if len(st.session_state.scenario_names) >= 5:
                st.warning("Please delete a scenario before adding (max 5).")
            elif not news_file:
                st.warning("Upload a .txt file first.")
            else:
                try:
                    text = news_file.read().decode("utf-8")[:12000]
                    client = get_openai_client()
                    if client is None:
                        raise RuntimeError("OpenAI client unavailable.")
                    prompt = (
                        "You are an energy market risk analyst. Read the news and propose stress scenario values "
                        "for the Philippine power market. Respond ONLY with a JSON object mapping the 14 variables: "
                        "duration (days), brent, jkm, coal, freight, insurance, logistics, fx, des_basis, "
                        "lng_availability, coal_delay, bid_uplift, reserve_alpha, bad_debt_multiplier. "
                        "Use reasonable stressed numbers. Example: "
                        '{"duration":30,"brent":85,"jkm":14,"coal":130,"freight":12,"insurance":1.0,'
                        '"logistics":1.0,"fx":59.5,"des_basis":1.0,"lng_availability":0.9,"coal_delay":7,'
                        '"bid_uplift":0.08,"reserve_alpha":0.3,"bad_debt_multiplier":1.1}'
                    )
                    completion = client.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[
                            {"role": "system", "content": prompt},
                            {"role": "user", "content": text},
                        ],
                        temperature=0.4,
                    )
                    reply = completion.choices[0].message.content.strip()
                    values = parse_json_safely(reply)
                    if not isinstance(values, dict):
                        raise ValueError(f"LLM did not return valid JSON. Raw reply: {reply[:200]}")
                    if add_scenario(new_scen_name, values):
                        st.success(f"Added scenario '{new_scen_name}' from news.")
                        st.rerun()
                except Exception as e:
                    st.error(f"Could not generate scenario: {e}")

        # Delete scenario control
        if len(st.session_state.scenario_names) > 1:
            del_name = st.selectbox(
                "Delete a scenario (cannot delete Base)",
                [n for n in st.session_state.scenario_names if n != st.session_state.scenario_names[0]],
                key="del_scenario_sel",
            )
            if st.button("Delete selected scenario", key="btn_del_scenario"):
                delete_scenario(del_name)
                st.rerun()

    # Tab 2 — Model Inputs
    with tab2:
        st.subheader("Model Inputs")
        c1, c2 = st.columns(2)
        st.session_state.coal_quality_key = c1.selectbox("Coal GAR case", list(defaults.COAL_QUALITY.keys()), index=1)
        st.session_state.season_key = c2.selectbox("Season", list(defaults.SEASONS.keys()), index=list(defaults.SEASONS.keys()).index(st.session_state.season_key))

        st.markdown("**Physical constants**")
        st.write(pd.DataFrame(defaults.CONSTANTS, index=["Value"]).T)

        with st.expander("Physical Constants & Coal Quality", expanded=False):
            st.write("Coal CV (GJ/t):", defaults.COAL_QUALITY[st.session_state.coal_quality_key]["gj_per_t"])

        with st.expander("Season & Demand", expanded=False):
            st.write("Seasonal multipliers")
            st.dataframe(pd.DataFrame(defaults.SEASONS).T)
            st.session_state.demand_df = st.data_editor(st.session_state.demand_df, num_rows="fixed")
            st.session_state.vre_df = st.data_editor(st.session_state.vre_df, num_rows="fixed", key="vre_editor")
            st.session_state.hydro_df = st.data_editor(st.session_state.hydro_df, num_rows="fixed", key="hydro_editor")

        with st.expander("SPC / WESM mitigation", expanded=False):
            st.session_state.spc_toggle = st.checkbox("SPC toggle", value=bool(st.session_state.spc_toggle))
            st.session_state.spc_cap = st.number_input("SPC cap (PHP/MWh)", value=float(st.session_state.spc_cap), step=100.0)

        with st.expander("Gen Co Plant Fleet", expanded=False):
            st.session_state.plants_df = st.data_editor(st.session_state.plants_df, num_rows="fixed")
            st.number_input("Oil product differential (USD/bbl)", value=float(defaults.PLANTS["oil_product_diff"]), disabled=True)

        with st.expander("Luzon WESM Stack", expanded=False):
            st.session_state.system_df = st.data_editor(st.session_state.system_df, num_rows="fixed")

        with st.expander("LNG Terminal Fees", expanded=False):
            st.session_state.lng_df = st.data_editor(st.session_state.lng_df, num_rows="fixed")

        with st.expander("Supply & Wholesale Contracts", expanded=False):
            st.session_state.contracts_df = st.data_editor(st.session_state.contracts_df, num_rows="fixed")

        with st.expander("Retail Portfolio", expanded=False):
            st.session_state.retail_df = st.data_editor(st.session_state.retail_df)

        with st.expander("Liquidity Position & Working Capital", expanded=False):
            st.session_state.liquidity_df = st.data_editor(st.session_state.liquidity_df)

        with st.expander("AGRA Regulatory Recovery", expanded=False):
            st.session_state.agra_df = st.data_editor(st.session_state.agra_df, num_rows="fixed")

        with st.expander("Trigger Playbook", expanded=False):
            st.session_state.triggers_df = st.data_editor(st.session_state.triggers_df)

    # Run engine
    inputs = build_inputs_from_state()
    results = engine.run_model(inputs)

    # Tab 3 — Results
    with tab3:
        st.subheader("Results & Dashboard")
        failed = [c for c in results["checks"] if not c["passed"]]
        if not failed:
            st.success("✓ MODEL OK — all checks pass")
        else:
            st.error("⚠️ Checks failing: " + ", ".join([c["name"] for c in failed]))

        st.markdown("**Section A: Scenario Matrix**")
        scen_df = scenario_df()
        st.dataframe(scen_df)
        ai_explain(
            "scenario",
            "scenario matrix",
            f"Scenarios: {st.session_state.scenario_names}. Key inputs:\n{df_to_text(scen_df)}",
        )

        st.markdown("**Section B: Generation Sensitivity**")
        metrics = ["Coal SRMC", "Gas SRMC", "Oil SRMC", "WESM Avg", "Reserve Price", "Generation Gross Margin"]
        units = ["PHP/MWh", "PHP/MWh", "PHP/MWh", "PHP/MWh", "PHP/MWh", "PHP mn/mo"]
        data = {"Metric": metrics, "Unit": units}
        for i, name in enumerate(st.session_state.scenario_names):
            data[name] = [
                results["plant_srmc"]["coal_srmc"][i],
                results["plant_srmc"]["gas_srmc"][i],
                results["plant_srmc"]["oil_srmc"][i],
                results["merit_order"]["wesm_avg"][i],
                results["reserve_pricing"]["reserve_price"][i],
                results["gen_margin"]["total_gen_margin"][i],
            ]
        table_b = pd.DataFrame(data)
        st.dataframe(table_b)
        c1, c2 = st.columns(2)
        c1.plotly_chart(chart_srmc(results, st.session_state.scenario_names), use_container_width=True)
        c2.plotly_chart(chart_wesm(results, st.session_state.scenario_names), use_container_width=True)
        ai_explain(
            "gen",
            "generation sensitivity",
            df_to_text(table_b),
        )

        st.markdown("**Section C: Supply & Retail Margin Sensitivity**")
        segs = ["BCQ Exposure", "Spot Sales", "Portfolio Retail EBIT", "Supply & Wholesale Margin"]
        data_c = {"Segment": segs, "Unit": ["PHP mn/mo"] * 4}
        for i, name in enumerate(st.session_state.scenario_names):
            data_c[name] = [
                results["supply_wholesale"]["bcq_margin"][i],
                results["supply_wholesale"]["spot_margin"][i],
                results["retail"]["portfolio_ebit"][i],
                results["supply_wholesale"]["total_sw_margin"][i],
            ]
        table_c = pd.DataFrame(data_c)
        st.dataframe(table_c)
        st.plotly_chart(chart_retail(results, st.session_state.scenario_names), use_container_width=True)
        ai_explain(
            "supply",
            "supply & retail",
            df_to_text(table_c),
        )

        st.markdown("**Section D: Liquidity Stress Structure**")
        liq = results["liquidity"]
        metrics_d = [
            "Working Capital",
            "LC Requirements",
            "Cash Balance",
            "Covenant Headroom",
            "Liquidity Runway",
            "Regulatory Receivable",
            "Immediate Recovery %",
            "Refund Provision",
        ]
        units_d = ["PHP mn", "PHP mn", "PHP mn", "PHP mn", "days", "PHP mn", "%", "PHP mn"]
        data_d = {"Metric": metrics_d, "Unit": units_d}
        for i, name in enumerate(st.session_state.scenario_names):
            data_d[name] = [
                liq["working_capital"][i],
                liq["lc_req"][i],
                liq["cash_balance"][i],
                liq["covenant_headroom"][i],
                liq["runway_days"][i],
                liq["reg_receivable"][i],
                results["liquidity"]["immediate_recovery"][i] * 100,
                liq["refund_provision"][i],
            ]
        liq_table = pd.DataFrame(data_d)
        st.dataframe(liq_table)
        st.plotly_chart(chart_liquidity(results, st.session_state.scenario_names), use_container_width=True)
        ai_explain(
            "liq",
            "liquidity",
            df_to_text(liq_table),
        )

        st.markdown("**Section E: Trigger Matrix**")
        trig_df = pd.DataFrame(results["triggers"])
        st.dataframe(trig_df)
        ai_explain(
            "triggers",
            "triggers",
            df_to_text(trig_df),
        )

        col1, col2 = st.columns(2)
        with col1:
            if st.button("📋 Copy to clipboard", use_container_width=True):
                st.code(scenario_df().to_csv(sep="\t", index=False))
        with col2:
            bio = download_results_xlsx(results, st.session_state.scenario_names)
            st.download_button(
                "⬇️ Download Results (.xlsx)",
                data=bio,
                file_name="GenCo_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    # Tab 4 — Intermediate calculations (summary tables only)
    with tab4:
        st.subheader("Intermediate Calculations")
        sub1, sub2, sub3, sub4, sub5, sub6, sub7, sub8, sub9 = st.tabs(
            [
                "Fuel Landing",
                "Plant SRMC",
                "Merit Order",
                "Reserve Pricing",
                "Gen Margin",
                "Wholesale Margin",
                "Margin Bridge",
                "Retail EBIT",
                "Liquidity Model",
            ]
        )

        with sub1:
            st.write(pd.DataFrame(results["fuel_landing"]))
        with sub2:
            st.write(pd.DataFrame({
                "Coal SRMC": results["plant_srmc"]["coal_srmc"],
                "Gas SRMC": results["plant_srmc"]["gas_srmc"],
                "Oil SRMC": results["plant_srmc"]["oil_srmc"],
                "Weighted fleet": results["plant_srmc"]["weighted_fleet_srmc"],
            }, index=st.session_state.scenario_names))
        with sub3:
            mo = results["merit_order"]
            st.write("Clearing prices (PHP/MWh)")
            st.dataframe(pd.DataFrame(mo["clearing"], index=["Off-peak", "Shoulder", "Peak"], columns=st.session_state.scenario_names))
        with sub4:
            st.write(pd.DataFrame(results["reserve_pricing"], index=st.session_state.scenario_names))
        with sub5:
            st.write(pd.DataFrame(results["gen_margin"], index=st.session_state.scenario_names))
        with sub6:
            st.write(pd.DataFrame(results["supply_wholesale"], index=st.session_state.scenario_names))
        with sub7:
            mb = pd.DataFrame(results["margin_bridge"]).T
            st.dataframe(mb)
            st.write("Reconciliation", results["margin_reconciliation"])
        with sub8:
            segs = results["retail"]["segments"]
            for seg in segs:
                st.write(seg["name"], pd.DataFrame(seg["ebit"], index=st.session_state.scenario_names, columns=["EBIT"]))
        with sub9:
            st.write(pd.DataFrame(results["liquidity"], index=st.session_state.scenario_names))


if __name__ == "__main__":
    main()
